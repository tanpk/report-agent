import pypdf
import os
import base64
import json
import hashlib
from google import genai
from google.genai import types
from dotenv import load_dotenv

load_dotenv()

# --- モデル設定 ---
MODEL_OCR = "gemini-2.5-flash"  # OCR処理

OCR_CACHE_DIR = ".ocr_cache"  # OCRキャッシュの保存先

# --- OCRキャッシュ ---

def _ocr_cache_key(file_path: str) -> str:
    """ファイルパス＋更新日時のハッシュをキーにする。ファイルが変われば再OCRする"""
    mtime = os.path.getmtime(file_path)
    raw = f"{os.path.abspath(file_path)}:{mtime}"
    return hashlib.sha256(raw.encode()).hexdigest()[:16]

def _load_ocr_cache(file_path: str) -> str | None:
    """OCRキャッシュが存在すれば読み込む"""
    os.makedirs(OCR_CACHE_DIR, exist_ok=True)
    key = _ocr_cache_key(file_path)
    cache_path = os.path.join(OCR_CACHE_DIR, f"{key}.json")
    if os.path.exists(cache_path):
        with open(cache_path, "r", encoding="utf-8") as f:
            print("  [OCRキャッシュ使用]")
            return json.load(f)["text"]
    return None

def _save_ocr_cache(file_path: str, text: str):
    """OCR結果をキャッシュに保存する"""
    os.makedirs(OCR_CACHE_DIR, exist_ok=True)
    key = _ocr_cache_key(file_path)
    cache_path = os.path.join(OCR_CACHE_DIR, f"{key}.json")
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump({"file": file_path, "text": text}, f, ensure_ascii=False)

# --- テキスト系の読み取り ---

def read_pdf(file_path: str) -> str:
    """テキスト埋め込みPDFからテキストを抽出する"""
    text = ""
    with open(file_path, "rb") as f:
        reader = pypdf.PdfReader(f)
        for page in reader.pages:
            text += page.extract_text()
    return text

def read_text(file_path: str) -> str:
    """テキストファイルを読み込む"""
    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()

# --- OCR系の読み取り ---

def _image_to_base64(image_path: str) -> tuple[str, str]:
    """画像ファイルをbase64に変換する。(base64文字列, MIMEタイプ) を返す"""
    ext = os.path.splitext(image_path)[1].lower()
    mime_map = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".webp": "image/webp",
    }
    mime_type = mime_map.get(ext, "image/jpeg")
    with open(image_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8"), mime_type

def _ocr_with_gemini(image_base64: str, mime_type: str) -> str:
    """Gemini APIで画像からテキストをOCRする"""
    client = genai.Client(api_key=os.getenv("GEMINI_API_KEY"))
    response = client.models.generate_content(
        model=MODEL_OCR,
        contents=[
            types.Part.from_bytes(
                data=base64.b64decode(image_base64),
                mime_type=mime_type,
            ),
            "この画像に含まれるテキストをすべて正確に抽出してください。レイアウトや改行も可能な限り再現してください。",
        ],
    )
    return response.text

def read_image(file_path: str) -> str:
    """画像ファイルをGemini APIでOCRしてテキストを返す。結果はキャッシュする"""
    cached = _load_ocr_cache(file_path)
    if cached:
        return cached

    image_base64, mime_type = _image_to_base64(file_path)
    result = _ocr_with_gemini(image_base64, mime_type)
    _save_ocr_cache(file_path, result)
    return result

def read_scanned_pdf(file_path: str) -> str:
    """スキャンPDF（画像PDF）を各ページOCRしてテキストを返す。結果はキャッシュする"""
    cached = _load_ocr_cache(file_path)
    if cached:
        return cached

    from pdf2image import convert_from_path
    import tempfile

    pages = convert_from_path(file_path, dpi=200)
    all_text = ""

    for i, page_image in enumerate(pages):
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            page_image.save(tmp.name, "PNG")
            tmp_path = tmp.name

        print(f"  OCR中: {i+1}/{len(pages)} ページ目...")
        # ページ画像はキャッシュせず（PDFごとまとめてキャッシュするため）
        image_base64, mime_type = _image_to_base64(tmp_path)
        page_text = _ocr_with_gemini(image_base64, mime_type)
        all_text += f"--- {i+1}ページ目 ---\n{page_text}\n\n"
        os.remove(tmp_path)

    _save_ocr_cache(file_path, all_text)
    return all_text

# --- Office系の読み取り ---

def read_docx(file_path: str) -> str:
    """Wordファイル（.docx）からテキストを抽出する"""
    from docx import Document
    doc = Document(file_path)
    return "\n".join(para.text for para in doc.paragraphs if para.text.strip())

def read_xlsx(file_path: str) -> str:
    """Excelファイル（.xlsx）の全シートをテキストに変換する"""
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    result = ""
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        result += f"--- シート：{sheet_name} ---\n"
        for row in ws.iter_rows(values_only=True):
            row_text = "\t".join("" if cell is None else str(cell) for cell in row)
            if row_text.strip():
                result += row_text + "\n"
        result += "\n"
    return result

# --- メインのルーター ---

def read_file(file_path: str, force_ocr: bool = False) -> str:
    """
    ファイルの種類を判別して読み込む。

    force_ocr=True にするとPDFをスキャン扱いでOCR処理する
    （テキスト抽出できないPDFに使う）
    """
    extension = os.path.splitext(file_path)[1].lower()

    if extension == ".pdf":
        if force_ocr:
            print("スキャンPDFとしてOCR処理します...")
            return read_scanned_pdf(file_path)
        else:
            text = read_pdf(file_path)
            if len(text.strip()) < 50:
                print("テキスト抽出量が少ないため、OCRに切り替えます...")
                return read_scanned_pdf(file_path)
            return text

    elif extension == ".txt":
        return read_text(file_path)

    elif extension in [".jpg", ".jpeg", ".png", ".gif", ".webp"]:
        return read_image(file_path)

    elif extension == ".docx":
        return read_docx(file_path)

    elif extension == ".xlsx":
        return read_xlsx(file_path)

    else:
        return f"非対応のファイル形式です：{extension}"
