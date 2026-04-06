import os
import json
import re
import openpyxl
from google import genai
from google.genai import types
from dotenv import load_dotenv

load_dotenv()

# --- モデル設定 ---
MODEL_LOW = "gemini-3.1-flash-lite-preview"  # 軽量処理（Excel列番号解析等）

MARKERS = ["*", "x", "o", "^", "square", "+", "diamond"]

# 近似曲線のプリセット
FIT_PRESETS = {
    "Logarithmic fit": {"expr": "-a*log(x)+b",  "params": ["a", "b"]},
    "Linear fit":      {"expr": "a*x+b",        "params": ["a", "b"]},
    "Power fit":       {"expr": "a*x.^n+b",     "params": ["a", "n", "b"]},
    "Exponential fit": {"expr": "a*exp(b*x)",   "params": ["a", "b"]},
}
LEGEND_LOCATIONS = ["southeast", "northeast", "northwest", "southwest", "best"]

SYSTEM_PROMPT = (
    "あなたはExcelデータを解析してMATLABグラフ生成を支援する専門家です。"
    "ユーザーの指定した列名をデータから探し、JSON形式のみで回答してください。"
    "余分な説明やMarkdownのコードブロックは不要です。"
)

class GraphAgent:

    def __init__(self):
        self.client = genai.Client(api_key=os.getenv("GEMINI_API_KEY"))

    # --- ① ユーザー入力収集 ---

    def collect_axes(self) -> dict:
        """横軸・縦軸のデータ名をユーザーに入力させる"""
        print("\n--- グラフの軸設定 ---")
        x_name = input("横軸のデータ名を入力してください（例: time）: ").strip()

        y_input = input("縦軸のデータ名を入力してください（複数はカンマ区切り、例: Ta, Tb, Tc）: ")
        y_names = [name.strip() for name in y_input.split(",")]

        x_label = input("横軸のラベル（例: τ: Elapsed Time）: ").strip()
        x_unit  = input("横軸の単位（例: min）: ").strip()
        y_label = input("縦軸のラベル（例: T: Temperature）: ").strip()
        y_unit  = input("縦軸の単位（例: ℃）: ").strip()

        print(f"凡例の位置の選択肢: {LEGEND_LOCATIONS}")
        legend_loc = input("凡例の位置（デフォルト: southeast）: ").strip()
        if legend_loc not in LEGEND_LOCATIONS:
            legend_loc = "southeast"

        y_min = input("Y軸の最小値（空欄で自動）: ").strip()
        y_max = input("Y軸の最大値（空欄で自動）: ").strip()
        ylim = None
        if y_min and y_max:
            ylim = [float(y_min), float(y_max)]

        fig_width  = input("図の幅px（デフォルト: 800）: ").strip() or "800"
        fig_height = input("図の高さpx（デフォルト: 600）: ").strip() or "600"

        return {
            "x_name": x_name,
            "y_names": y_names,
            "x_label": x_label,
            "x_unit": x_unit,
            "y_label": y_label,
            "y_unit": y_unit,
            "legend_location": legend_loc,
            "ylim": ylim,
            "fig_width": int(fig_width),
            "fig_height": int(fig_height),
        }

    # --- ② xlsxをテキスト化してGeminiで解析 ---

    def analyze_xlsx(self, xlsx_path: str, axes: dict) -> dict:
        """
        xlsxをテキスト化してGeminiに渡し、列番号のみを解析させる。
        戻り値例：
        {
            "x_col_idx": 3,
            "y_col_idxs": [4, 5, 6, 7, 8],
        }
        """
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        xlsx_text = ""
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            xlsx_text += f"=== シート: {sheet_name} ===\n"
            # 列番号ヘッダーを先頭行に追加（空列があっても列番号がずれないようにする）
            col_header = "\t".join(f"[col{i}]" for i in range(1, ws.max_column + 1))
            xlsx_text += col_header + "\n"
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join("" if v is None else str(v) for v in row)
                if row_text.strip():
                    xlsx_text += row_text + "\n"
            xlsx_text += "\n"

        x = axes["x_name"]
        ys = ", ".join(axes["y_names"])

        prompt = (
            "以下のExcelデータを解析してください。\n\n"
            f"{xlsx_text}\n"
            f'探す列：\n- 横軸: "{x}"\n- 縦軸: {ys}\n\n'
            "以下の条件で解析してください：\n"
            "1. 横軸と縦軸が両方含まれているシート・表を特定する\n"
            "2. 同じ名称が複数の表に登場する場合は、横軸と縦軸が共存する表を優先する\n"
            "3. 列番号はExcel全体で左から何番目か（1始まり）を返す\n\n"
            "以下のJSON形式のみで回答してください（説明文不要）：\n"
            "{\n"
            '  "x_col_idx": 横軸の列番号,\n'
            '  "y_col_idxs": [縦軸1の列番号, 縦軸2の列番号, ...]\n'
            "}"
        )

        response = self.client.models.generate_content(
            model=MODEL_LOW,
            contents=prompt,
            config=types.GenerateContentConfig(
                system_instruction=SYSTEM_PROMPT,
                max_output_tokens=256,
            )
        )

        text = response.text.strip()
        # コードブロックを除去
        text = re.sub(r"```json|```", "", text).strip()
        # JSONブロックを正規表現で抽出（余分なテキストが混入する場合の保険）
        match = re.search(r'{.*?}', text, re.DOTALL)
        if not match:
            raise ValueError(f"Geminiの返答からJSONを抽出できませんでした: {text}")
        text = match.group(0)
        return json.loads(text)

    # --- ③ .matファイル生成 ---

    def save_mat(self, xlsx_path: str, axes: dict, analysis: dict, mat_filename: str):
        """
        xlsxから該当列を抽出してscipy.io.savematで.matファイルを生成する。
        変数名はユーザー入力（axes）のものを使う。
        """
        import scipy.io
        import numpy as np

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active

        # 数値データ行のみ抽出
        data_rows = [
            row for row in ws.iter_rows(values_only=True)
            if any(isinstance(v, (int, float)) for v in row)
        ]

        x_idx = analysis["x_col_idx"] - 1  # 0始まりに変換
        y_idxs = [i - 1 for i in analysis["y_col_idxs"]]

        mat_data = {}
        mat_data[axes["x_name"]] = np.array(
            [row[x_idx] for row in data_rows], dtype=float
        )
        for y_name, y_idx in zip(axes["y_names"], y_idxs):
            mat_data[y_name] = np.array(
                [row[y_idx] if row[y_idx] is not None else float("nan")
                 for row in data_rows], dtype=float
            )

        scipy.io.savemat(mat_filename, mat_data)
        print(f"{mat_filename} を生成しました。")

    # --- ④ MATLABコード生成（APIなし） ---

    def generate_matlab(self, axes: dict, mat_filename: str) -> str:
        """
        .matを読み込んでグラフを描画するMATLABコードを生成する。
        """
        x_name     = axes["x_name"]
        y_names    = axes["y_names"]
        # 軸ラベル: 変数名をイタリック体に ($x$ 形式)
        x_var      = axes["x_label"].split(":")[0].strip()
        x_rest     = axes["x_label"][len(x_var):]
        x_label    = f"${x_var}${x_rest} [{axes['x_unit']}]"
        y_var      = axes["y_label"].split(":")[0].strip()
        y_rest     = axes["y_label"][len(y_var):]
        y_label    = f"${y_var}${y_rest} [{axes['y_unit']}]"

        xlim       = axes.get("xlim")
        ylim       = axes.get("ylim")
        x_scale    = axes.get("x_scale", "linear")
        y_scale    = axes.get("y_scale", "linear")
        fig_width  = axes.get("fig_width", 800)
        fig_height = axes.get("fig_height", 600)
        show_legend = axes.get("show_legend", True)
        show_grid   = axes.get("show_grid", False)
        png_name    = axes.get("png_name", "graph.png")
        fit_curves  = axes.get("fit_curves", [])

        lines = []
        lines.append(f"load {mat_filename}")
        lines.append("")
        lines.append(f"fig = figure('Position', [100, 100, {fig_width}, {fig_height}]);")

        # データのプロット
        lines.append(f'plot({x_name}, {y_names[0]}, "{MARKERS[0]}")')
        lines.append(f'xlabel("{x_label}")')
        lines.append(f'ylabel("{y_label}")')

        needs_hold = len(y_names) > 1 or bool(fit_curves)
        if needs_hold:
            lines.append("hold on")
            for i, y_name in enumerate(y_names[1:], 1):
                lines.append(f'plot({x_name}, {y_name}, "{MARKERS[i % len(MARKERS)]}")')

        # 近似曲線（実線/破線を自動判定）
        fit_legend_labels = []
        for fit in fit_curves:
            fit_lines, fit_label = self._build_fit_curve_block(fit, x_name, xlim)
            lines.extend(fit_lines)
            fit_legend_labels.append(fit_label)

        # 凡例
        if show_legend:
            legend_labels = ", ".join(f"'Data'" if len(y_names) == 1 else f"'{y}'" for y in y_names)
            if len(y_names) > 1:
                legend_labels = ", ".join(f"'{y}'" for y in y_names)
            else:
                legend_labels = "'Data'"
            if fit_legend_labels:
                legend_labels += ", " + ", ".join(f"'{l}'" for l in fit_legend_labels)
            lines.append(f"legend({legend_labels}, 'Location', '{axes['legend_location']}')")

        if needs_hold:
            lines.append("hold off")

        # 軸範囲・スケール
        if xlim:
            lines.append(f"xlim([{xlim[0]} {xlim[1]}]);")
        if ylim:
            lines.append(f"ylim([{ylim[0]} {ylim[1]}]);")
        if x_scale == "log":
            lines.append("set(gca, 'XScale', 'log');")
        if y_scale == "log":
            lines.append("set(gca, 'YScale', 'log');")

        # グリッド線
        if show_grid:
            lines.append("grid on;")
            lines.append("set(gca, 'GridColor', [0.5 0.5 0.5], 'GridAlpha', 0.5);")

        lines.append("")
        lines.append("set(fig, 'Color', 'white');")
        lines.append("set(gca, 'Color', 'white');")
        lines.append("set(gca, 'XColor', 'black', 'YColor', 'black');")
        if show_legend:
            lines.append("leg = legend;")
            lines.append("set(leg, 'Color', 'white', 'TextColor', 'black', 'EdgeColor', 'black');")
        lines.append(f"saveas(fig, '{png_name}');")
        lines.append("exit;")

        return "\n".join(lines)

    def _build_fit_curve_block(self, fit: dict, x_col: str, xlim: list | None = None) -> tuple[list[str], str]:
        """
        近似曲線のMATLABコードブロックを生成する。
        データ範囲内は実線、外挿範囲は破線で描画する。
        戻り値: (コード行リスト, 凡例ラベル)
        """
        import re
        lines = []
        for name, value in fit["params"].items():
            lines.append(f"{name} = {value};")

        # x範囲: xlimが指定されていればその全域、なければデータ範囲
        if xlim:
            x_start, x_end = xlim[0], xlim[1]
        else:
            x_start, x_end = f"min({x_col})", f"max({x_col})"

        # データ範囲（実線）と外挿範囲（破線）を分けて描画
        # データ範囲: min(x_col)〜max(x_col)
        expr_template = re.sub(r'(?<![a-zA-Z_])x(?![a-zA-Z_0-9])', '{VAR}', fit["expr"])

        # 外挿範囲（データより左）
        if xlim:
            lines.append(f"x_pre = linspace({x_start}, min({x_col}), 200);")
            lines.append(f"f_pre = {expr_template.replace('{VAR}', 'x_pre')};")
            lines.append("plot(x_pre, f_pre, '--', 'Color', [0.5 0.5 0.5])")

        # データ範囲内（実線）
        lines.append(f"x_fit = linspace(min({x_col}), max({x_col}), 500);")
        lines.append(f"f_fit = {expr_template.replace('{VAR}', 'x_fit')};")
        lines.append('plot(x_fit, f_fit, "-")')

        # 外挿範囲（データより右）
        if xlim:
            lines.append(f"x_post = linspace(max({x_col}), {x_end}, 200);")
            lines.append(f"f_post = {expr_template.replace('{VAR}', 'x_post')};")
            lines.append("plot(x_post, f_post, '--', 'Color', [0.5 0.5 0.5])")

        label = fit.get("label", "Fit")
        return lines, label

    # --- 保存 ---

    def save_matlab(self, code: str, filename: str):
        """生成したMATLABコードを.mファイルに保存する"""
        with open(filename, "w", encoding="utf-8") as f:
            f.write(code)
        print(f"{filename} に保存しました。")
